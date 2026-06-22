# -*- coding: utf-8 -*-
"""
ناصع — مُولّد تقرير الصيانة اللايف
Nasie — Live Maintenance Report Builder

الاستخدام / Usage:
    python build_report.py <input.xlsx> <output_dir>

يقرأ ملف إكسل (ورقتا "Service Calls" و"Sheet1") ويُخرج:
    <output_dir>/index.html   (لوحة التقرير الخضراء الحيّة)

Reads an Excel file (sheets "Service Calls" + "Sheet1") and writes:
    <output_dir>/index.html   (the live green dashboard)

يحفظ لقطة الفترة الحالية في prev_summary.json لاستخدامها في المقارنة بالفترة التالية.
Saves a snapshot to prev_summary.json for next run's comparison.
"""
import sys, os, re, json, math
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook

# ============== اللوغو المضمّن / embedded logo ==============
LOGO_B64 = open(os.path.join(os.path.dirname(__file__), "logo.b64")).read().strip()

AR_DAYS = {0:'الاثنين',1:'الثلاثاء',2:'الأربعاء',3:'الخميس',4:'الجمعة',5:'السبت',6:'الأحد'}
AR_MONTHS = {1:'يناير',2:'فبراير',3:'مارس',4:'أبريل',5:'مايو',6:'يونيو',7:'يوليو',8:'أغسطس',9:'سبتمبر',10:'أكتوبر',11:'نوفمبر',12:'ديسمبر'}
CAT_ICON={'صيانة دورية':'🔄','تكييف':'❄️','إنترنت':'🌐','أبواب ونوافذ':'🚪','كهرباء':'⚡',
 'دهانات':'🎨','صيانة عامة':'🔧','سباكة':'🚿','تشطيبات':'🏗️','خدمات توصيل':'📦',
 'بطاريات':'🔋','أقفال ذكية':'🔑','مواقف سيارات':'🅿️','أثاث':'🪑'}
CITY_ICON={'الرياض':'🏙️','المدينة':'🕌'}
PAL=['#6EEBA3','#38bdf8','#34d399','#22d3ee','#4ade80','#60a5fa','#2dd4bf','#a3e635','#10b981','#0ea5e9']
EXCLUDE_HINTS=['demo listing','general requests - internal','inquiry','استفسار','guest experience']

def map_city(c, area):
    s=((c or '')+' '+(area or '')).lower()
    if 'madinah' in s or 'madina' in s: return 'المدينة'
    return 'الرياض'

def map_cat(cat2, cat3):
    a=(cat2 or '').lower(); b=(cat3 or '').lower(); s=a+' '+b
    if 'paint' in a or 'walls' in b: return 'دهانات'
    if 'internet' in s: return 'إنترنت'
    if 'electric' in s: return 'كهرباء'
    if 'a/c' in s or ' ac' in s or 'air cond' in s: return 'تكييف'
    if 'door' in s or 'window' in s: return 'أبواب ونوافذ'
    if 'plumb' in s or 'blockage' in s or 'water leak' in s: return 'سباكة'
    if 'batter' in s: return 'بطاريات'
    if 'fit-out' in s or 'fitout' in s: return 'تشطيبات'
    if 'deliver' in s: return 'خدمات توصيل'
    if 'parking' in s: return 'مواقف سيارات'
    if 'smart' in s or 'lock' in s or ' key' in s: return 'أقفال ذكية'
    if 'furniture' in s: return 'أثاث'
    if 'preventive' in a: return 'صيانة دورية'
    if 'general maintenance' in b: return 'صيانة عامة'
    if 'maintenance' in a: return 'صيانة دورية'
    return 'صيانة دورية'

def excluded(text):
    t=(text or '').lower()
    return any(h in t for h in EXCLUDE_HINTS)

def load_data(xlsx_path):
    wb=load_workbook(xlsx_path, data_only=True)
    names=wb.sheetnames
    SC={}
    if 'Service Calls' in names:
        for r in list(wb['Service Calls'].iter_rows(values_only=True))[1:]:
            if r[0] is None: continue
            d=r[7]
            if isinstance(d,str):
                try: d=datetime.strptime(d[:10],'%Y-%m-%d')
                except: d=None
            SC[str(r[0])]=dict(listing=r[2],desc=r[3],status=r[4],date=d,contractor=r[6],
                               mat=float(r[9] or 0),labor=float(r[10] or 0))
    S1={}
    sheet1 = 'Sheet1' if 'Sheet1' in names else names[-1]
    for r in list(wb[sheet1].iter_rows(values_only=True))[1:]:
        url=str(r[16] or ''); m=re.search(r'(\d{6,})',url)
        if not m: continue
        S1[m.group(1)]=dict(cat2=(r[1] or '').strip(),cat3=(r[2] or '').strip(),
            contractor=(r[3] or '').strip(),listing=(r[4] or '').strip(),city=(r[6] or '').strip(),
            area=(r[7] or '').strip(),internal=float(r[9] or 0),external=float(r[10] or 0),
            extflag=(r[11] or ''),admin=float(r[12] or 0),material=float(r[13] or 0),
            total=float(r[14] or 0),raiser=(r[17] or '').strip())
    # Build the request set. Prefer Service Calls (real dates+status); enrich from Sheet1.
    recs=[]
    src = SC if SC else {tid:dict(listing=b['listing'],desc=b['cat2'],status='Done',date=None,
                                  contractor=b['contractor'],mat=b['material'],labor=b['internal']) for tid,b in S1.items()}
    for tid,sv in src.items():
        b=S1.get(tid)
        if excluded((b['cat2'] if b else '')+' '+(b['cat3'] if b else '')+' '+(sv.get('desc') or '')):
            continue
        status=(sv.get('status') or 'Done')
        # only DONE / In Review (= Pending Approval)
        if status not in ('Done','In Review','Pending Approval'): continue
        if b:
            internal,external,admin,material=b['internal'],b['external'],b['admin'],b['material']
            cat=map_cat(b['cat2'],b['cat3']); city=map_city(b['city'],b['area'])
            raiser=b['raiser'] or 'AI'; listing=(b['listing'] or sv.get('listing') or '').strip()
            desc=b['cat2'] or sv.get('desc')
        else:
            internal=float(sv.get('labor') or 0); external=0.0; admin=0.0; material=float(sv.get('mat') or 0)
            cat=map_cat(sv.get('desc'),sv.get('desc')); city=map_city(None,sv.get('listing'))
            raiser='AI'; listing=(sv.get('listing') or '').strip(); desc=sv.get('desc')
        if not raiser or raiser.lower() in ('','none','غير محدد'): raiser='AI'
        d=sv.get('date') or datetime(2000,1,1)
        total=internal+external+admin+material
        recs.append(dict(id=tid,listing=listing or '—',cat=cat,city=city,
            internal=round(internal,2),external=round(external,2),admin=round(admin,2),
            material=round(material,2),total=round(total,2),raiser=raiser,contractor=sv.get('contractor') or '',
            status=status,date=d.strftime('%Y-%m-%d'),dow=AR_DAYS[d.weekday()],day=d.day,month=d.month,desc=desc))
    recs.sort(key=lambda x:(x['date'],x['id']))
    return recs

def f(x):
    x=round(float(x),2)
    return f"{x:,.0f}" if abs(x-round(x))<0.005 else f"{x:,.2f}"
def pct(x,base): return f"{(100.0*x/base):.1f}%" if base else "0%"

def build_html(recs, prev):
    N=len(recs); TOT=round(sum(r['total'] for r in recs),2)
    def S(k): return round(sum(r[k] for r in recs),2)
    sm=dict(n=N,labor=round(S('internal')+S('external'),2),internal=S('internal'),external=S('external'),
            material=S('material'),admin=S('admin'),total=TOT)
    sm['avg']=round(TOT/N,2) if N else 0
    top=max(recs,key=lambda x:x['total']) if recs else dict(total=0,id='',listing='—')
    sm['max']=top['total']; sm['max_id']=top['id']; sm['max_listing']=top['listing']
    ext_n=sum(1 for r in recs if r['external']>0); noext_n=N-ext_n
    dates=sorted(set(r['date'] for r in recs)) or [datetime.now().strftime('%Y-%m-%d')]
    d0=datetime.strptime(dates[0],'%Y-%m-%d'); d1=datetime.strptime(dates[-1],'%Y-%m-%d')
    period=f"{d0.day} {AR_MONTHS[d0.month]} إلى {d1.day} {AR_MONTHS[d1.month]} {d1.year}"
    period_short=f"{d0.day}–{d1.day} {AR_MONTHS[d1.month]}"

    def grp(keyf):
        g=defaultdict(lambda:dict(n=0,internal=0,external=0,admin=0,material=0,total=0,recs=[]))
        for r in recs:
            x=g[keyf(r)]; x['n']+=1
            for k in ('internal','external','admin','material','total'): x[k]=round(x[k]+r[k],2)
            x['recs'].append(r)
        return g
    cat_rows=sorted(grp(lambda r:r['cat']).items(),key=lambda kv:(-kv[1]['n'],-kv[1]['total']))
    rais_rows=sorted(grp(lambda r:r['raiser']).items(),key=lambda kv:-kv[1]['n'])
    city_rows=sorted(grp(lambda r:r['city']).items(),key=lambda kv:-kv[1]['n'])
    day_rows=sorted(grp(lambda r:r['date']).items())
    top10=sorted(recs,key=lambda r:-r['total'])[:10]
    mats=sorted([r for r in recs if r['material']>0],key=lambda r:-r['material'])
    extdet=[r for r in recs if r['external']>0]

    def cu(v,dec=False): return f'<span class="cu" data-c="{v}" data-d="{1 if dec else 0}">0</span>'
    def kbox(v,l,c='mint',ic=''):
        return f'<div class="kpi {c}"><div class="kic">{ic}</div><div class="kv">{v}</div><div class="kl">{l}</div></div>'
    def donut(segs,center='',sub='',rounded=False,size=190,stroke=24):
        total=sum(v for _,v,_ in segs) or 1; r=(size-stroke)/2; cx=cy=size/2; circ=2*math.pi*r
        cer=' stroke-linecap="round"' if rounded else ''
        o=f'<div class="dn"><svg viewBox="0 0 {size} {size}" class="dns"><circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="rgba(255,255,255,.05)" stroke-width="{stroke}"/>'
        off=0
        for lab,val,col in segs:
            dash=val/total*circ; rot=off/total*360-90
            o+=f'<circle class="seg" cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{col}" stroke-width="{stroke}" stroke-dasharray="{dash:.2f} {circ:.2f}" stroke-dashoffset="{dash:.2f}" data-off="0" transform="rotate({rot:.2f} {cx} {cy})"{cer}/>'
            off+=val
        return o+f'</svg><div class="dnc"><div class="dnv">{center}</div><div class="dnl">{sub}</div></div></div>'
    def legend(segs,total,unit=''):
        o='<div class="lg">'
        for lab,val,col in segs:
            v=f(val)+((' '+unit) if unit else '')
            o+=f'<div class="lgi"><span class="dot" style="background:{col}"></span><span class="lgn">{lab}</span><span class="lgv">{v} · {pct(val,total)}</span></div>'
        return o+'</div>'
    def hbar(label,val,base,color,disp):
        w=100.0*val/base if base else 0; inside=w>=14; p=pct(val,base)
        if inside:
            fill=f'<div class="bf" data-w="{w:.1f}" style="background:linear-gradient(90deg,{color},{color}cc)"><span class="bfp in">{p}</span></div>'
        else:
            fill=f'<div class="bf" data-w="{max(w,2):.1f}" style="background:linear-gradient(90deg,{color},{color}cc)"></div><span class="bfp out" style="color:{color}">{p}</span>'
        return f'<div class="bh"><div class="bl">{label}</div><div class="bt">{fill}</div><div class="bv">{disp}</div></div>'
    def idl(t): return f'<a class="tl" href="https://app.boomnow.com/dashboard/listing-task/{t}" target="_blank">#{t}</a>' if t else '—'

    parts=[]
    # 0 SUMMARY
    kpis=[(N and cu(N) or '0','إجمالي الطلبات','mint','🧾'),('SAR '+cu(TOT,1),'إجمالي التكلفة','blue','💰'),
          ('SAR '+cu(sm['avg'],1),'متوسط تكلفة الطلب','mint','📊'),('SAR '+cu(sm['labor'],1),'تكلفة العمالة','blue','👷'),
          ('SAR '+cu(sm['material'],1),'إجمالي المواد','mint','🧰'),('SAR '+cu(sm['admin'],1),'رسوم الإدارة','blue','🏢')]
    kg=''.join(kbox(v,l,c,ic) for v,l,c,ic in kpis)
    cseg=[('العمالة',sm['labor'],'#6EEBA3'),('رسوم الإدارة',sm['admin'],'#38bdf8'),('المواد',sm['material'],'#22d3ee')]
    topcat=cat_rows[0] if cat_rows else ('—',{'n':0,'total':0})
    parts.append(f"""<div class="card"><div class="ch2"><span class="ci">📊</span> الملخص العام</div><div class="kgrid">{kg}</div></div>
<div class="row"><div class="card flex1"><div class="ch3">توزيع التكلفة</div>{donut(cseg,'SAR '+f(TOT),'إجمالي')}{legend(cseg,TOT,'SAR')}</div>
<div class="card flex1"><div class="ch3">🔍 أبرز المؤشرات</div>
<div class="insi"><span class="ic2">{CAT_ICON.get(topcat[0],'🔧')}</span><div><b>{topcat[1]['n']} طلب — {topcat[0]}</b> ({pct(topcat[1]['n'],N)}) أعلى تصنيف عدداً بقيمة SAR {f(topcat[1]['total'])}.</div></div>
<div class="insi"><span class="ic2">💎</span><div>أعلى طلب {idl(sm['max_id'])} — {sm['max_listing']} — <b>SAR {f(sm['max'])}</b> ({pct(sm['max'],TOT)} من التكلفة).</div></div>
<div class="insi"><span class="ic2">👷</span><div><b>{pct(noext_n,N)}</b> اعتماد داخلي — {noext_n} طلب بدون فني خارجي مقابل {ext_n} بفني خارجي.</div></div>
</div></div>
<div class="card hilite"><span class="ic2">💎</span> أعلى تكلفة طلب: {idl(sm['max_id'])} — {sm['max_listing']} — <b>SAR {f(sm['max'])}</b></div>""")
    # 1 COSTS
    seg=[('الفني الداخلي',sm['internal'],'#6EEBA3'),('الفني الخارجي',sm['external'],'#38bdf8'),('رسوم الإدارة',sm['admin'],'#2dd4bf'),('المواد',sm['material'],'#60a5fa')]
    rws=[('إجمالي تكلفة العمالة',sm['labor'],False),('— الفني الخارجي',sm['external'],True),('— الفني الداخلي',sm['internal'],True),('إجمالي المواد',sm['material'],False),('إجمالي رسوم الإدارة',sm['admin'],False)]
    trs=''.join(f'<tr><td class="{"sub" if s else "b"}">{n}</td><td class="am">{f(v)}</td><td>{pct(v,TOT)}</td></tr>' for n,v,s in rws)
    parts.append(f"""<div class="row"><div class="card flex1"><div class="ch3">تركيبة التكاليف</div>{donut(seg,'SAR '+f(TOT),'الإجمالي')}{legend(seg,TOT,'SAR')}</div>
<div class="card flex1"><div class="ch3">التفصيل</div><table><thead><tr><th>البيان</th><th>المبلغ</th><th>النسبة</th></tr></thead><tbody>{trs}
<tr class="tot"><td>الإجمالي الكلي</td><td class="am">{f(TOT)}</td><td>100%</td></tr></tbody></table>
<div style="margin-top:14px">{hbar('العمالة',sm['labor'],TOT,'#6EEBA3','SAR '+f(sm['labor']))}{hbar('الإدارة',sm['admin'],TOT,'#38bdf8','SAR '+f(sm['admin']))}{hbar('المواد',sm['material'],TOT,'#22d3ee','SAR '+f(sm['material']))}</div></div></div>""")
    # 2 DISTRIBUTION
    parts.append(f"""<div class="row"><div class="card flex1"><div class="ch3">نوع التنفيذ</div>{donut([('بدون فني خارجي',noext_n,'#6EEBA3'),('مع فني خارجي',ext_n,'#38bdf8')],cu(N),'طلب',True)}{legend([('بدون فني خارجي',noext_n,'#6EEBA3'),('مع فني خارجي',ext_n,'#38bdf8')],N,'طلب')}</div>
<div class="card flex1"><div class="ch3">توزيع الطلبات</div><div class="kgrid2">{kbox(cu(noext_n),'بدون فني خارجي — '+pct(noext_n,N),'mint')}{kbox(cu(ext_n),'مع فني خارجي — '+pct(ext_n,N),'blue')}</div>
<div style="margin-top:16px">{hbar('بدون فني خارجي',noext_n,N,'#6EEBA3',f(noext_n)+' طلب')}{hbar('مع فني خارجي',ext_n,N,'#38bdf8',f(ext_n)+(' طلب' if ext_n==1 else ' طلبات'))}</div></div></div>""")
    # 3 CATEGORIES
    bars=''.join(hbar(f"{CAT_ICON.get(k,'🔧')} {k}",v['n'],N,PAL[i%len(PAL)],f(v['n'])+' طلب') for i,(k,v) in enumerate(cat_rows))
    trs=''.join(f"<tr><td>{CAT_ICON.get(k,'🔧')} {k}</td><td>{v['n']}</td><td>{pct(v['n'],N)}</td><td class='am'>{f(v['external'])}</td><td class='am'>{f(v['internal'])}</td><td class='am'>{f(v['material'])}</td><td class='am'>{f(v['admin'])}</td><td class='am b'>{f(v['total'])}</td></tr>" for k,v in cat_rows)
    trs+=f"<tr class='tot'><td>الإجمالي</td><td>{N}</td><td>100%</td><td class='am'>{f(sm['external'])}</td><td class='am'>{f(sm['internal'])}</td><td class='am'>{f(sm['material'])}</td><td class='am'>{f(sm['admin'])}</td><td class='am'>{f(TOT)}</td></tr>"
    det=''
    for k,v in cat_rows:
        rr=''.join(f"<tr><td>{idl(r['id'])}</td><td>{r['listing']}</td><td class='am'>{f(r['internal'])}</td><td class='am'>{f(r['external'])}</td><td class='am'>{f(r['material'])}</td><td class='am'>{f(r['admin'])}</td><td class='am b'>{f(r['total'])}</td></tr>" for r in v['recs'])
        det+=f"""<div class="acc" onclick="tc(this)"><span>{CAT_ICON.get(k,'🔧')} {k} <span class="pill">{v['n']} طلب</span> <span class="pill blue">SAR {f(v['total'])}</span></span><span class="ar">▾</span></div><div class="acb"><table><thead><tr><th>رقم الطلب</th><th>الشقة</th><th>داخلي</th><th>خارجي</th><th>مواد</th><th>إدارة</th><th>إجمالي</th></tr></thead><tbody>{rr}</tbody></table></div>"""
    parts.append(f"""<div class="card"><div class="ch2"><span class="ci">🏷️</span> تحليل التصنيفات</div>{bars}</div>
<div class="card"><div class="ch3">جدول التصنيفات</div><table><thead><tr><th>التصنيف</th><th>العدد</th><th>النسبة</th><th>خارجي</th><th>داخلي</th><th>مواد</th><th>إدارة</th><th>إجمالي</th></tr></thead><tbody>{trs}</tbody></table></div>
<div class="card"><div class="ch3">تفاصيل كل تصنيف</div>{det}</div>""")
    # 4 RAISERS
    bars=''.join(hbar(k,v['n'],N,PAL[i%len(PAL)],f(v['n'])+(' طلب' if v['n']==1 else ' طلبات')) for i,(k,v) in enumerate(rais_rows))
    trs=''.join(f"<tr><td>{k}</td><td>{v['n']}</td><td>{pct(v['n'],N)}</td><td class='am'>{f(v['internal']+v['external'])}</td><td class='am'>{f(v['material'])}</td><td class='am'>{f(v['admin'])}</td><td class='am b'>{f(v['total'])}</td></tr>" for k,v in rais_rows)
    trs+=f"<tr class='tot'><td>الإجمالي</td><td>{N}</td><td>100%</td><td class='am'>{f(sm['labor'])}</td><td class='am'>{f(sm['material'])}</td><td class='am'>{f(sm['admin'])}</td><td class='am'>{f(TOT)}</td></tr>"
    parts.append(f"""<div class="card"><div class="ch2"><span class="ci">👤</span> رافع الطلبات</div>{bars}</div>
<div class="card"><div class="ch3">جدول رافعي الطلبات</div><table><thead><tr><th>المُنشئ</th><th>العدد</th><th>النسبة</th><th>العمالة</th><th>مواد</th><th>إدارة</th><th>إجمالي</th></tr></thead><tbody>{trs}</tbody></table></div>""")
    # 5 GEO
    gseg=[(CITY_ICON.get(k,'')+' '+k,v['total'],PAL[i]) for i,(k,v) in enumerate(city_rows)]
    trs=''.join(f"<tr><td>{CITY_ICON.get(k,'')} {k}</td><td>{v['n']}</td><td>{pct(v['n'],N)}</td><td class='am'>{f(v['internal'])}</td><td class='am'>{f(v['external'])}</td><td class='am'>{f(v['material'])}</td><td class='am'>{f(v['admin'])}</td><td class='am b'>{f(v['total'])}</td></tr>" for k,v in city_rows)
    trs+=f"<tr class='tot'><td>الإجمالي</td><td>{N}</td><td>100%</td><td class='am'>{f(sm['internal'])}</td><td class='am'>{f(sm['external'])}</td><td class='am'>{f(sm['material'])}</td><td class='am'>{f(sm['admin'])}</td><td class='am'>{f(TOT)}</td></tr>"
    parts.append(f"""<div class="row"><div class="card flex1"><div class="ch3">توزيع التكلفة جغرافياً</div>{donut(gseg,'SAR '+f(TOT),'إجمالي')}{legend(gseg,TOT,'SAR')}</div>
<div class="card flex1"><div class="ch3">حسب المدينة</div><div style="margin-top:6px">{''.join(hbar(CITY_ICON.get(k,'')+' '+k,v['total'],TOT,PAL[i],'SAR '+f(v['total'])) for i,(k,v) in enumerate(city_rows))}</div>
<table style="margin-top:10px"><thead><tr><th>المدينة</th><th>عدد</th><th>نسبة</th><th>داخلي</th><th>خارجي</th><th>مواد</th><th>إدارة</th><th>إجمالي</th></tr></thead><tbody>{trs}</tbody></table></div></div>""")
    # 6 MATERIALS
    trs=''.join(f"<tr><td>{i}</td><td>{r['desc']}</td><td>{idl(r['id'])}</td><td>{r['listing']}</td><td>{r['day']} {AR_MONTHS.get(r['month'],'')}</td><td class='am b'>{f(r['material'])}</td></tr>" for i,r in enumerate(mats,1))
    ms=sum(r['material'] for r in mats); trs+=f"<tr class='tot'><td colspan='5'>الإجمالي</td><td class='am'>{f(ms)}</td></tr>"
    parts.append(f"""<div class="card"><div class="ch2"><span class="ci">🔧</span> المواد</div><table><thead><tr><th>#</th><th>الوصف</th><th>رقم الطلب</th><th>الشقة</th><th>التاريخ</th><th>التكلفة</th></tr></thead><tbody>{trs}</tbody></table>
<div class="note">💡 يحتسب هذا القسم تكلفة المواد فقط (Material Cost). عدد الطلبات التي تضمّنت مواد: {len(mats)} بإجمالي SAR {f(ms)}.</div></div>""")
    # 7 TECHNICIANS
    trs=''.join(f"<tr><td>{idl(r['id'])}</td><td>{r['listing']}</td><td>{r['desc']}</td><td class='am'>{f(r['internal'])}</td><td class='am'>{f(r['external'])}</td><td class='am'>{f(r['admin'])}</td><td class='am b'>{f(r['total'])}</td></tr>" for r in sorted(extdet,key=lambda x:-x['external']))
    ti=sum(r['internal'] for r in extdet);te=sum(r['external'] for r in extdet);ta=sum(r['admin'] for r in extdet);tt=sum(r['total'] for r in extdet)
    trs+=f"<tr class='tot'><td colspan='3'>الإجمالي</td><td class='am'>{f(ti)}</td><td class='am'>{f(te)}</td><td class='am'>{f(ta)}</td><td class='am'>{f(tt)}</td></tr>"
    cb=''
    for k,v in city_rows:
        en=sum(1 for r in v['recs'] if r['external']>0)
        cb+=f"<tr><td>{CITY_ICON.get(k,'')} {k}</td><td class='am'>{f(v['internal'])}</td><td class='am'>{f(v['external'])}</td><td>{en}</td></tr>"
    cb+=f"<tr class='tot'><td>الإجمالي</td><td class='am'>{f(sm['internal'])}</td><td class='am'>{f(sm['external'])}</td><td>{ext_n}</td></tr>"
    parts.append(f"""<div class="row"><div class="card flex1"><div class="ch3">الداخلي مقابل الخارجي</div>{donut([('داخلي',sm['internal'],'#6EEBA3'),('خارجي',sm['external'],'#38bdf8')],'SAR '+f(sm['labor']),'العمالة')}{legend([('الفني الداخلي',sm['internal'],'#6EEBA3'),('الفني الخارجي',sm['external'],'#38bdf8')],sm['labor'],'SAR')}</div>
<div class="card flex1"><div class="ch3">ملخص</div><div class="kgrid2">{kbox('SAR '+cu(sm['internal'],1),'الفني الداخلي — '+str(noext_n)+' طلب','mint')}{kbox('SAR '+cu(sm['external'],1),'الفني الخارجي — '+str(ext_n)+' طلبات','blue')}</div>
<div class="note" style="margin-top:14px">الفني الخارجي يمثل {pct(sm['external'],sm['labor'])} من إجمالي العمالة، مركّز في الدهان والتكييف.</div></div></div>
<div class="card"><div class="ch3">تفاصيل طلبات الفني الخارجي</div><table><thead><tr><th>رقم الطلب</th><th>الشقة</th><th>الوصف</th><th>داخلي</th><th>خارجي</th><th>إدارة</th><th>إجمالي</th></tr></thead><tbody>{trs}</tbody></table></div>
<div class="card"><div class="ch3">التوزيع حسب المدينة</div><table><thead><tr><th>المدينة</th><th>داخلي</th><th>خارجي</th><th>عدد خارجي</th></tr></thead><tbody>{cb}</tbody></table></div>""")
    # 8 TOP10
    mx=top10[0]['total'] if top10 else 1
    trs=''.join(f"<tr><td><span class='rank'>{i}</span></td><td>{idl(r['id'])}</td><td>{r['listing']}</td><td class='am'>{f(r['internal'])}</td><td class='am'>{f(r['external'])}</td><td class='am'>{f(r['material'])}</td><td class='am'>{f(r['admin'])}</td><td class='am b'>{f(r['total'])}</td><td style='width:120px'><div class='mbar'><div class='mbf' data-w='{100*r['total']/mx:.1f}'></div></div></td></tr>" for i,r in enumerate(top10,1))
    t3=sum(r['total'] for r in top10[:3])
    parts.append(f"""<div class="card"><div class="ch2"><span class="ci">📈</span> أعلى 10 طلبات تكلفة</div><table><thead><tr><th>#</th><th>رقم الطلب</th><th>الشقة</th><th>داخلي</th><th>خارجي</th><th>مواد</th><th>إدارة</th><th>إجمالي</th><th>المؤشر</th></tr></thead><tbody>{trs}</tbody></table>
<div class="note">📈 أعلى 3 طلبات تستحوذ على SAR {f(t3)} أي {pct(t3,TOT)} من إجمالي الأسبوع.</div></div>""")
    # 9 DAILY
    dmax=max((v['total'] for _,v in day_rows),default=1) or 1
    cols=''
    for dt,v in day_rows:
        r0=v['recs'][0]
        cols+=f"<div class='col'><div class='colbarwrap'><div class='colval'>SAR {f(v['total'])}</div><div class='colbar' data-h='{max(100*v['total']/dmax,3):.1f}'></div></div><div class='colx'>{r0['dow']}<br><span>{r0['day']}/{r0['month']}</span></div><div class='coln'>{v['n']} طلب</div></div>"
    trs=''.join(f"<tr><td>{v['recs'][0]['dow']}</td><td>{v['recs'][0]['day']} {AR_MONTHS.get(v['recs'][0]['month'],'')}</td><td>{v['n']}</td><td class='am'>{f(v['internal'])}</td><td class='am'>{f(v['external'])}</td><td class='am'>{f(v['material'])}</td><td class='am'>{f(v['admin'])}</td><td class='am b'>{f(v['total'])}</td></tr>" for dt,v in day_rows)
    trs+=f"<tr class='tot'><td colspan='2'>الإجمالي</td><td>{N}</td><td class='am'>{f(sm['internal'])}</td><td class='am'>{f(sm['external'])}</td><td class='am'>{f(sm['material'])}</td><td class='am'>{f(sm['admin'])}</td><td class='am'>{f(TOT)}</td></tr>"
    det=''
    for dt,v in day_rows:
        r0=v['recs'][0]
        rr=''.join(f"<tr><td>{idl(r['id'])}</td><td>{r['listing']}</td><td>{CAT_ICON.get(r['cat'],'')} {r['cat']}</td><td class='am'>{f(r['internal'])}</td><td class='am'>{f(r['external'])}</td><td class='am'>{f(r['material'])}</td><td class='am'>{f(r['admin'])}</td><td class='am b'>{f(r['total'])}</td></tr>" for r in v['recs'])
        det+=f"""<div class="acc" onclick="tc(this)"><span>📅 {r0['dow']} {r0['day']} {AR_MONTHS.get(r0['month'],'')} <span class="pill">{v['n']} طلب</span> <span class="pill blue">SAR {f(v['total'])}</span></span><span class="ar">▾</span></div><div class="acb"><table><thead><tr><th>رقم الطلب</th><th>الشقة</th><th>التصنيف</th><th>داخلي</th><th>خارجي</th><th>مواد</th><th>إدارة</th><th>إجمالي</th></tr></thead><tbody>{rr}</tbody></table></div>"""
    parts.append(f"""<div class="card"><div class="ch2"><span class="ci">📅</span> الطلبات اليومية</div><div class="chart-cols">{cols}</div></div>
<div class="card"><div class="ch3">الجدول اليومي</div><table><thead><tr><th>اليوم</th><th>التاريخ</th><th>العدد</th><th>داخلي</th><th>خارجي</th><th>مواد</th><th>إدارة</th><th>إجمالي</th></tr></thead><tbody>{trs}</tbody></table></div>
<div class="card"><div class="ch3">تفاصيل كل يوم</div>{det}</div>""")
    # 10 COMPARISON (vs prev snapshot)
    if prev:
        def dr(label,cur,pv):
            dv=cur-pv; up=dv>0; arr='▲' if up else('▼' if dv<0 else '—'); cls='up' if up else('down' if dv<0 else 'flat')
            chg=f"{abs(100.0*dv/pv):.1f}%" if pv else '—'; sign='+' if up else ''
            return f"<tr><td>{label}</td><td class='am'>{f(cur)}</td><td class='am'>{f(pv)}</td><td class='am'>{sign}{f(dv)}</td><td><span class='tag {cls}'>{arr} {chg}</span></td></tr>"
        crows=(dr('عدد الطلبات',N,prev['n'])+dr('تكلفة العمالة',sm['labor'],prev['labor'])+dr('الفني الخارجي',sm['external'],prev['external'])
        +dr('الفني الداخلي',sm['internal'],prev['internal'])+dr('المواد',sm['material'],prev['material'])+dr('رسوم الإدارة',sm['admin'],prev['admin'])+dr('الإجمالي',TOT,prev['total']))
        comp=f"""<div class="card"><div class="ch2"><span class="ci">📊</span> المقارنة بالفترة السابقة</div>
<div class="kgrid2">{kbox('SAR '+cu(TOT,1),'الحالي ('+period_short+') — '+str(N)+' طلب','blue')}{kbox('SAR '+f(prev['total']),'السابق ('+prev.get('period','—')+') — '+str(prev['n'])+' طلب','dim')}</div>
<table style="margin-top:14px"><thead><tr><th>البيان</th><th>الحالي</th><th>السابق</th><th>الفرق</th><th>التغيير</th></tr></thead><tbody>{crows}</tbody></table></div>"""
    else:
        comp=f"""<div class="card"><div class="ch2"><span class="ci">📊</span> المقارنة بالفترة السابقة</div>
<div class="note">لا توجد بيانات فترة سابقة بعد. ستظهر المقارنة تلقائياً ابتداءً من الفترة القادمة (يُحفظ ملخص هذه الفترة كمرجع).</div></div>"""
    parts.append(comp)
    # 11 RECOMMENDATIONS (auto from data)
    tc3=cat_rows[:3]
    recos=[('1️⃣','متابعة أعلى التصنيفات',f"تصنيف «{tc3[0][0]}» هو الأعلى ({tc3[0][1]['n']} طلب) — يستحق متابعة الأسباب الجذرية." if tc3 else '—'),
           ('2️⃣','مراجعة أعلى طلب تكلفة',f"الطلب #{sm['max_id']} بقيمة SAR {f(sm['max'])} يمثل {pct(sm['max'],TOT)} من التكلفة — تأكد من مبرراته."),
           ('3️⃣','اعتماد الطلبات قيد المراجعة',f"{sum(1 for r in recs if r['status'] in ('In Review','Pending Approval'))} طلب قيد المراجعة — يُنصح بتسريع الاعتماد."),
           ('4️⃣','توثيق المواد',f"{len(mats)} طلب فقط سجّل مواد — تدقيق إدخال البيانات يحسّن دقة التكاليف."),
           ('5️⃣','الاستمرار بالصيانة الوقائية','البرنامج الوقائي يقلّل الأعطال الكبرى ويضبط التكاليف.')]
    rc=''.join(f'<div class="reco"><div class="rn">{n}</div><div><div class="rt">{t}</div><div class="rd">{x}</div></div></div>' for n,t,x in recos)
    parts.append(f"""<div class="card"><div class="ch2"><span class="ci">💡</span> الملخص النهائي والتوصيات</div>
<div class="insi"><span class="ic2">📋</span><div><b>إجمالي {N} طلب بقيمة SAR {f(TOT)}</b> خلال {period}.</div></div>
<div class="insi"><span class="ic2">🏙️</span><div><b>المدن:</b> {' · '.join(CITY_ICON.get(k,'')+' '+k+' '+str(v['n'])+' طلب' for k,v in city_rows)}.</div></div>
<div class="insi"><span class="ic2">✅</span><div><b>{pct(noext_n,N)} اعتماد داخلي</b> — تكلفة الخارجي SAR {f(sm['external'])} ({pct(sm['external'],sm['labor'])} من العمالة).</div></div></div>
<div class="card"><div class="ch3">💡 التوصيات</div>{rc}</div>""")

    TABS=['الملخص','التكاليف','التوزيع','التصنيفات','رافع الطلبات','الجغرافي','المواد','الفنيين','الأداء','اليومي','المقارنة','التوصيات']
    TABIC=['📊','💰','📈','🏷️','👤','🗺️','🔧','👷','📈','📅','📊','💡']
    tabbtns=''.join(f'<button class="tab{" on" if i==0 else ""}" onclick="st({i})">{TABIC[i]} {t}</button>' for i,t in enumerate(TABS))
    tps=''.join(f'<div class="tp{" on" if i==0 else ""}" id="t{i}">{p}</div>' for i,p in enumerate(parts))
    html=PAGE_TMPL.format(CSS=CSS,LOGO=LOGO_B64,N=N,TOT=f(TOT),PERIOD=period,TABS=tabbtns,BODY=tps,SCRIPT=SCRIPT,
                          GEN=datetime.now().strftime('%Y-%m-%d %H:%M'))
    snapshot=dict(n=N,labor=sm['labor'],internal=sm['internal'],external=sm['external'],
                  material=sm['material'],admin=sm['admin'],total=TOT,period=period_short)
    return html, snapshot

# ====== CSS / TEMPLATE / SCRIPT (مطابقة للوحة المعتمدة) ======
CSS = r"""<style>
@import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;800;900&display=swap');
*{margin:0;padding:0;box-sizing:border-box}
:root{--mint:#6EEBA3;--blue:#38bdf8;--ink:#eafff5;--tx:#c3e2d3;--mut:#80a293;--bd:rgba(110,235,163,.16);--up:#34d399;--down:#38bdf8}
body{font-family:'Tajawal',sans-serif;color:var(--tx);direction:rtl;font-size:13px;line-height:1.65;background:#06100c;background-image:radial-gradient(900px 500px at 85% -5%,rgba(56,189,248,.14),transparent 60%),radial-gradient(900px 600px at 5% 0%,rgba(110,235,163,.16),transparent 55%),radial-gradient(700px 700px at 50% 120%,rgba(45,212,191,.10),transparent 60%);background-attachment:fixed;min-height:100vh}
.ct{max-width:1220px;margin:0 auto;padding:22px}
.rh{background:linear-gradient(135deg,rgba(15,61,46,.92),rgba(6,22,17,.92));border:1px solid var(--bd);border-radius:22px;padding:26px 34px;margin-bottom:18px;position:relative;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.45),inset 0 1px 0 rgba(255,255,255,.05)}
.rh::after{content:'';position:absolute;inset:0;background:radial-gradient(600px 200px at 80% 0,rgba(110,235,163,.18),transparent 70%)}
.rhx{display:flex;align-items:center;justify-content:space-between;gap:20px;position:relative;z-index:1}
.rht h1{font-size:27px;font-weight:900;color:#fff;letter-spacing:-.5px}.rht .sub{font-size:14px;color:var(--mut);margin-top:3px}
.rhl{height:78px;filter:drop-shadow(0 6px 18px rgba(110,235,163,.35))}
.bdg{display:inline-block;background:linear-gradient(135deg,var(--mint),#34d399);color:#06231a;padding:5px 15px;border-radius:30px;font-weight:800;font-size:12px;margin:10px 6px 0 0;box-shadow:0 6px 18px rgba(110,235,163,.3)}
.bdg.b{background:linear-gradient(135deg,var(--blue),#0ea5e9);color:#04222e;box-shadow:0 6px 18px rgba(56,189,248,.3)}
.tc{position:sticky;top:10px;z-index:99;margin-bottom:18px}
.tabs{display:flex;flex-wrap:wrap;gap:6px;padding:9px;background:rgba(8,20,16,.75);backdrop-filter:blur(14px);border:1px solid var(--bd);border-radius:16px;box-shadow:0 12px 36px rgba(0,0,0,.4)}
.tab{padding:8px 14px;border:none;background:0 0;cursor:pointer;border-radius:10px;font-family:inherit;font-size:12px;font-weight:600;color:var(--mut);transition:.25s;white-space:nowrap}
.tab:hover{background:rgba(110,235,163,.1);color:var(--ink)}.tab.on{background:linear-gradient(135deg,var(--mint),#2dd4bf);color:#06231a;font-weight:800;box-shadow:0 6px 18px rgba(110,235,163,.35)}
.tp{display:none;animation:fade .5s ease}.tp.on{display:block}@keyframes fade{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:none}}
.card{background:linear-gradient(180deg,rgba(18,38,30,.66),rgba(10,24,19,.6));backdrop-filter:blur(12px);border:1px solid var(--bd);border-radius:18px;padding:22px;margin-bottom:16px;box-shadow:0 14px 40px rgba(0,0,0,.35),inset 0 1px 0 rgba(255,255,255,.04)}
.row{display:flex;gap:16px;flex-wrap:wrap;align-items:stretch}.flex1{flex:1;min-width:300px;margin-bottom:0}.row>.card{margin-bottom:16px}
.ch2{font-size:19px;font-weight:900;color:#fff;margin-bottom:18px;display:flex;align-items:center;gap:10px}
.ci{display:inline-grid;place-items:center;width:38px;height:38px;border-radius:11px;background:linear-gradient(135deg,rgba(110,235,163,.22),rgba(56,189,248,.18));border:1px solid var(--bd);font-size:18px}
.ch3{font-size:15px;font-weight:800;color:var(--mint);margin-bottom:14px;padding-right:12px;border-right:3px solid var(--mint)}
.kgrid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.kgrid2{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}
.kpi{position:relative;border-radius:16px;padding:18px;overflow:hidden;border:1px solid var(--bd);background:linear-gradient(160deg,rgba(255,255,255,.03),rgba(255,255,255,0));transition:transform .25s,box-shadow .25s}
.kpi:hover{transform:translateY(-4px)}.kpi.mint{box-shadow:0 10px 30px rgba(110,235,163,.12);background:linear-gradient(160deg,rgba(110,235,163,.16),rgba(110,235,163,.02))}
.kpi.blue{box-shadow:0 10px 30px rgba(56,189,248,.12);background:linear-gradient(160deg,rgba(56,189,248,.16),rgba(56,189,248,.02))}.kpi.dim{background:linear-gradient(160deg,rgba(255,255,255,.04),transparent)}
.kic{font-size:20px;margin-bottom:8px;opacity:.9}.kv{font-size:27px;font-weight:900;color:#fff;font-variant-numeric:tabular-nums;line-height:1.1}.kl{font-size:11.5px;color:var(--mut);margin-top:5px}
table{width:100%;border-collapse:separate;border-spacing:0;font-size:12px}
th{background:rgba(110,235,163,.1);color:var(--mint);padding:11px 12px;text-align:right;font-weight:800;white-space:nowrap;border-bottom:1px solid var(--bd)}
th:first-child{border-radius:0 10px 0 0}th:last-child{border-radius:10px 0 0 0}
td{padding:9px 12px;border-bottom:1px solid rgba(255,255,255,.05);text-align:right;font-variant-numeric:tabular-nums;color:var(--tx)}
tbody tr{transition:background .15s}tbody tr:hover td{background:rgba(110,235,163,.05)}
.am{text-align:left;direction:ltr}.b{font-weight:800;color:#fff}.sub{padding-right:26px;color:var(--mut)}
.tot td{background:linear-gradient(90deg,rgba(110,235,163,.14),rgba(56,189,248,.1))!important;font-weight:900;color:#fff;border-top:1px solid var(--mint)}
.dn{position:relative;width:200px;height:200px;margin:6px auto 10px}.dns{width:200px;height:200px}
.seg{transition:stroke-dashoffset 1.1s cubic-bezier(.22,1,.36,1)}
.dnc{position:absolute;inset:0;display:grid;place-content:center;text-align:center}.dnv{font-size:20px;font-weight:900;color:#fff}.dnl{font-size:11px;color:var(--mut)}
.lg{display:flex;flex-direction:column;gap:8px}.lgi{display:flex;align-items:center;gap:9px;font-size:12.5px}
.dot{width:12px;height:12px;border-radius:4px;flex-shrink:0;box-shadow:0 0 10px currentColor}.lgn{flex:1;color:var(--tx)}.lgv{color:var(--mut);font-variant-numeric:tabular-nums}
.bh{display:flex;align-items:center;gap:10px;margin-bottom:9px}.bl{min-width:140px;font-weight:600;font-size:12.5px;color:var(--tx)}
.bt{flex:1;height:26px;background:rgba(255,255,255,.05);border-radius:13px;border:1px solid rgba(255,255,255,.04);position:relative}
.bf{height:100%;width:0;border-radius:13px;display:flex;align-items:center;justify-content:flex-start;padding:0 10px;transition:width 1.1s cubic-bezier(.22,1,.36,1);box-shadow:0 0 16px rgba(110,235,163,.25);white-space:nowrap;max-width:100%}
.bfp{font-size:11px;font-weight:800;font-variant-numeric:tabular-nums}.bfp.in{color:#04231a}.bfp.out{position:absolute;left:10px;top:50%;transform:translateY(-50%)}
.bv{min-width:96px;text-align:left;font-weight:800;font-size:12px;color:#fff;font-variant-numeric:tabular-nums}
.insi{display:flex;gap:12px;padding:12px 14px;background:rgba(110,235,163,.06);border:1px solid var(--bd);border-radius:12px;margin-bottom:9px;font-size:13px;line-height:1.7}
.insi.down{background:rgba(56,189,248,.07);border-color:rgba(56,189,248,.2)}.insi b{color:#fff}.ic2{font-size:18px}
.hilite{display:flex;align-items:center;gap:10px;font-size:14px;background:linear-gradient(90deg,rgba(110,235,163,.12),rgba(56,189,248,.08));font-weight:600}.hilite b{color:var(--mint)}
.note{background:rgba(56,189,248,.07);border-right:3px solid var(--blue);padding:12px 15px;border-radius:10px;margin-top:14px;font-size:12.5px;color:var(--tx)}
a.tl{color:var(--blue);text-decoration:none;font-weight:700}a.tl:hover{text-decoration:underline}
.acc{cursor:pointer;padding:13px 16px;background:rgba(255,255,255,.03);border:1px solid var(--bd);border-radius:12px;display:flex;justify-content:space-between;align-items:center;margin-bottom:7px;font-weight:700;color:var(--ink);transition:.2s}
.acc:hover{background:rgba(110,235,163,.08)}.acb{display:none;padding:6px 6px 14px;animation:fade .4s}.acb.op{display:block}.ar{transition:transform .3s;color:var(--mint)}.ar.op{transform:rotate(180deg)}
.pill{display:inline-block;background:rgba(110,235,163,.16);color:var(--mint);padding:2px 10px;border-radius:20px;font-size:11px;font-weight:800;margin-right:4px}.pill.blue{background:rgba(56,189,248,.16);color:var(--blue)}
.tag{display:inline-block;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:800;font-variant-numeric:tabular-nums}
.tag.up{background:rgba(52,211,153,.16);color:#34d399}.tag.down{background:rgba(56,189,248,.16);color:#38bdf8}.tag.flat{background:rgba(255,255,255,.08);color:var(--mut)}
.rank{display:inline-grid;place-items:center;width:26px;height:26px;border-radius:8px;background:linear-gradient(135deg,var(--mint),#2dd4bf);color:#06231a;font-weight:900;font-size:12px}
.mbar{height:8px;background:rgba(255,255,255,.06);border-radius:6px;overflow:hidden}.mbf{height:100%;width:0;background:linear-gradient(90deg,var(--mint),var(--blue));border-radius:6px;transition:width 1.1s cubic-bezier(.22,1,.36,1)}
.chart-cols{display:flex;align-items:flex-end;gap:10px;height:280px;padding:10px 4px 0}
.col{flex:1;display:flex;flex-direction:column;align-items:center;height:100%}.colbarwrap{flex:1;width:100%;display:flex;flex-direction:column;justify-content:flex-end;align-items:center}
.colbar{width:62%;max-width:52px;height:0;border-radius:9px 9px 0 0;background:linear-gradient(180deg,var(--mint),#15916a);box-shadow:0 0 20px rgba(110,235,163,.3);transition:height 1.1s cubic-bezier(.22,1,.36,1)}
.colval{font-size:10.5px;color:var(--ink);font-weight:800;margin-bottom:5px;opacity:0;transition:opacity .5s .6s;white-space:nowrap}.col.show .colval{opacity:1}
.colx{font-size:11px;color:var(--tx);margin-top:8px;text-align:center;font-weight:600}.colx span{color:var(--mut);font-size:10px}.coln{font-size:10px;color:var(--mint);margin-top:2px;font-weight:700}
.reco{display:flex;gap:14px;padding:14px;background:rgba(255,255,255,.03);border:1px solid var(--bd);border-radius:13px;margin-bottom:10px}
.rn{font-size:18px;flex-shrink:0}.rt{font-weight:800;color:#fff;font-size:13.5px}.rd{color:var(--mut);font-size:12.5px;margin-top:2px}
.foot{text-align:center;color:var(--mut);font-size:11px;padding:14px;opacity:.7}
@media(max-width:640px){.kgrid{grid-template-columns:1fr}.kgrid2{grid-template-columns:1fr}.bl{min-width:96px}}
@media print{body{background:#06100c!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}.tc,.foot{display:none!important}.tp{display:block!important;page-break-before:always}#t0{page-break-before:avoid!important}.row{display:block}.flex1{margin-bottom:12px}.card{break-inside:avoid;page-break-inside:avoid}.card:has(table),.card:has(.acc){break-inside:auto}table{break-inside:auto}thead{display:table-header-group}tr{break-inside:avoid}.seg{stroke-dashoffset:0!important}.bf,.mbf,.colbar{transition:none!important}.colval{opacity:1!important}.acb{display:block!important}.ar{display:none}}
@page{size:A4;margin:10mm}
</style>"""

SCRIPT = r"""<script>
function fmt(v,dec){return dec?v.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}):Math.round(v).toLocaleString('en-US')}
function animTab(p){if(!p)return;
 p.querySelectorAll('.cu').forEach(function(e){if(e.dataset.done)return;e.dataset.done=1;var t=parseFloat(e.dataset.c),dec=e.dataset.d=='1',s=performance.now(),D=900;function fr(n){var k=Math.min((n-s)/D,1),v=t*(1-Math.pow(1-k,3));e.textContent=fmt(v,dec);if(k<1)requestAnimationFrame(fr);else e.textContent=fmt(t,dec)}requestAnimationFrame(fr)});
 setTimeout(function(){p.querySelectorAll('.bf').forEach(function(e){e.style.width=e.dataset.w+'%'});p.querySelectorAll('.seg').forEach(function(e){e.style.strokeDashoffset=e.dataset.off});p.querySelectorAll('.mbf').forEach(function(e){e.style.width=e.dataset.w+'%'});p.querySelectorAll('.colbar').forEach(function(e,i){setTimeout(function(){e.style.height=e.dataset.h+'%';e.closest('.col').classList.add('show')},i*70)})},60)}
function st(i){document.querySelectorAll('.tp').forEach(t=>t.classList.remove('on'));document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));var p=document.getElementById('t'+i);p.classList.add('on');document.querySelectorAll('.tab')[i].classList.add('on');window.scrollTo({top:0,behavior:'smooth'});animTab(p)}
function tc(h){var c=h.nextElementSibling,a=h.querySelector('.ar');c.classList.toggle('op');if(a)a.classList.toggle('op')}
window.addEventListener('load',function(){animTab(document.getElementById('t0'))});
window.addEventListener('beforeprint',function(){document.querySelectorAll('.cu').forEach(e=>e.textContent=fmt(parseFloat(e.dataset.c),e.dataset.d=='1'));document.querySelectorAll('.bf,.mbf').forEach(e=>e.style.width=e.dataset.w+'%');document.querySelectorAll('.seg').forEach(e=>e.style.strokeDashoffset='0');document.querySelectorAll('.colbar').forEach(e=>e.style.height=e.dataset.h+'%')});
</script>"""

PAGE_TMPL = """<!DOCTYPE html><html lang="ar" dir="rtl"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>تقرير الصيانة — ناصع — {PERIOD}</title>{CSS}</head><body><div class="ct">
<div class="rh"><div class="rhx"><div class="rht"><h1>تقرير الصيانة الأسبوعي</h1>
<div class="sub">الفترة {PERIOD} — ناصع لخدمات الصيانة</div>
<span class="bdg">{N} طلب صيانة</span><span class="bdg b">SAR {TOT}</span></div>
<img class="rhl" src="data:image/png;base64,{LOGO}" alt="ناصع"></div></div>
<div class="tc"><div class="tabs">{TABS}</div></div>
{BODY}
<div class="foot">ناصع لخدمات الصيانة · تقرير آلي · آخر تحديث {GEN}</div>
</div>{SCRIPT}</body></html>"""

def main():
    if len(sys.argv)<3:
        print("Usage: python build_report.py <input.xlsx> <output_dir>"); sys.exit(1)
    xlsx, outdir = sys.argv[1], sys.argv[2]
    os.makedirs(outdir, exist_ok=True)
    prev_path=os.path.join(outdir,'prev_summary.json')
    prev=None
    if os.path.exists(prev_path):
        try: prev=json.load(open(prev_path,encoding='utf-8'))
        except: prev=None
    recs=load_data(xlsx)
    if not recs:
        print("WARNING: no qualifying records found."); 
    html, snap = build_html(recs, prev)
    open(os.path.join(outdir,'index.html'),'w',encoding='utf-8').write(html)
    json.dump(snap, open(prev_path,'w',encoding='utf-8'), ensure_ascii=False, indent=1)
    print(f"OK — wrote {os.path.join(outdir,'index.html')} ({len(recs)} records, total SAR {snap['total']:,.2f})")

if __name__=='__main__':
    main()
